"""
Synchronise Class-Hierarchy-V1.csv depuis l'Excel original des encadrantes.
Garde le format du CSV existant (colonnes, séparateur ;) en ajoutant Sub-class 5.
"""
import openpyxl
import csv
import pathlib

EXCEL = pathlib.Path(r'C:\Users\core solutions\Desktop\Stage\Ressources\Last_year_stagiaire_work\Class_Hierarchy_VF_09avr26xlsx.xlsx')
CSV_OUT = pathlib.Path(__file__).parent / 'data-csv' / 'Class-Hierarchy-V1.csv'
CSV_BACKUP = pathlib.Path(__file__).parent / 'data-csv' / 'Class-Hierarchy-V1-backup.csv'

# Sauvegarder l'ancien CSV
import shutil
shutil.copy(CSV_OUT, CSV_BACKUP)
print(f'Backup: {CSV_BACKUP}')

# Lire l'Excel
wb = openpyxl.load_workbook(EXCEL)
ws = wb['Hierarchy']
xl_headers = [ws.cell(1, c).value or '' for c in range(1, ws.max_column + 1)]

# Index des colonnes Excel qui nous intéressent
cols = {
    'CLASS':       xl_headers.index('CLASS'),
    'sub-class 1': xl_headers.index('Sub-class 1'),
    'sub-class 2': xl_headers.index('Sub-class 2'),
    'sub-class 3': xl_headers.index('Sub-class 3'),
    'sub-class 4': xl_headers.index('Sub-class 4'),
    'sub-class 5': xl_headers.index('Sub-class 5'),
    'Synonym':     xl_headers.index('Synonym'),
    'Definition':  xl_headers.index('Definition'),
    'Outils':      xl_headers.index('Validated tools'),
}

# Lire l'ancien CSV pour récupérer ses colonnes supplémentaires (stats, etc.)
with open(CSV_BACKUP, encoding='utf-8-sig', newline='') as f:
    reader = csv.reader(f, delimiter=';')
    old_headers = next(reader)
    old_rows = {
        (r[0].strip(), r[1].strip() if len(r) > 1 else '',
         r[2].strip() if len(r) > 2 else '', r[3].strip() if len(r) > 3 else ''): r
        for r in reader
    }

# Construire les nouveaux headers (CLASS + sub-class 1→5 + reste de l'ancien CSV)
new_headers = ['CLASS', 'sub-class 1', 'sub-class 2', 'sub-class 3', 'sub-class 4', 'sub-class 5']
# Ajouter les colonnes de l'ancien CSV (Synonym et après)
extra_start = old_headers.index('Synonym') if 'Synonym' in old_headers else 5
extra_cols = old_headers[extra_start:]
new_headers += extra_cols
n_extra = len(extra_cols)

# Générer les nouvelles lignes depuis l'Excel
new_rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    cl  = str(row[cols['CLASS']] or '').strip()
    s1  = str(row[cols['sub-class 1']] or '').strip()
    s2  = str(row[cols['sub-class 2']] or '').strip()
    s3  = str(row[cols['sub-class 3']] or '').strip()
    s4  = str(row[cols['sub-class 4']] or '').strip()
    s5  = str(row[cols['sub-class 5']] or '').strip()
    syn = str(row[cols['Synonym']] or '').strip()

    if not cl:
        continue

    # Chercher les données supplémentaires dans l'ancien CSV
    key = (cl, s1, s2, s3)
    old = old_rows.get(key, None)
    if old:
        extra = old[extra_start:extra_start + n_extra]
        # Compléter si trop court
        extra = list(extra) + [''] * (n_extra - len(extra))
    else:
        extra = [syn] + [''] * (n_extra - 1)

    new_rows.append([cl, s1, s2, s3, s4, s5] + extra[:n_extra])

# Écrire le nouveau CSV
with open(CSV_OUT, 'w', encoding='utf-8', newline='') as f:
    writer = csv.writer(f, delimiter=';')
    writer.writerow(new_headers)
    writer.writerows(new_rows)

print(f'Ancien CSV : {len(old_rows)} lignes')
print(f'Nouveau CSV : {len(new_rows)} lignes')

# Vérifier les concepts
from collections import Counter
class_cols = [0, 1, 2, 3, 4, 5]
concepts = set()
for r in new_rows:
    for i in class_cols:
        if i < len(r) and r[i].strip():
            concepts.add(r[i].strip().lower())
print(f'Concepts uniques dans nouveau CSV : {len(concepts)}')
print(f'Backup sauvegardé : {CSV_BACKUP}')
